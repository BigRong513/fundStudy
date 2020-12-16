# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import openpyxl

class FundDataPipeline:
    manager_table_name = 'manager.xlsx'
    manager_sheet_name = 'list'
    data_table_path = r'E:/project/fundStudy/data/'

    def open_spider(self, spider):
        print("Here pipeline open spider")
        self.manager_table = openpyxl.load_workbook(self.data_table_path+self.manager_table_name)
        self.manager_sheet = self.manager_table[self.manager_sheet_name]

        self.manager_sheet.cell(1, 1).value = '姓名'
        self.manager_sheet.cell(1, 2).value = '公司'
        self.manager_sheet.cell(1, 3).value = '从业时间'
        self.manager_sheet.cell(1, 4).value = '现任基金资产总规模'
        self.manager_sheet.cell(1, 5).value = '现任基金最佳回报'
        self.manager_sheet.cell(1, 6).value = '基金代码'
        self.manager_sheet.cell(1, 7).value = '基金名称'
        self.manager_sheet.cell(1, 8).value = '基金类型'
        self.manager_sheet.cell(1, 9).value = '规模(亿元)'
        self.manager_sheet.cell(1, 10).value = '任职时间'
        self.manager_sheet.cell(1, 11).value = '任职天数'
        self.manager_sheet.cell(1, 12).value = '任职回报'
        self.current_row = 2

    def close_spider(self, spider):
        print("Here pipeline close spider")
        self.manager_table.save(self.data_table_path+self.manager_table_name)
        self.manager_table.close()


    def process_item(self, item, spider):
        self.manager_sheet.cell(self.current_row, 1).value = item['name']
        self.manager_sheet.cell(self.current_row, 2).value = item['company']
        self.manager_sheet.cell(self.current_row, 3).value = item['work_time']
        self.manager_sheet.cell(self.current_row, 4).value = item['scope']
        self.manager_sheet.cell(self.current_row, 5).value = item['best_return']
        self.manager_sheet.cell(self.current_row, 6).value = item['fund_code']
        self.manager_sheet.cell(self.current_row, 7).value = item['fund_name']
        self.manager_sheet.cell(self.current_row, 8).value = item['fund_type']
        self.manager_sheet.cell(self.current_row, 9).value = item['fund_scope']
        self.manager_sheet.cell(self.current_row, 10).value = item['fund_time']
        self.manager_sheet.cell(self.current_row, 11).value = item['fund_day']
        self.manager_sheet.cell(self.current_row, 12).value = item['fund_return']
        self.current_row += 1
        return item
